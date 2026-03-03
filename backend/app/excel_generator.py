import os
import re
from typing import List, Optional
from openpyxl import load_workbook

from app.models import TechProcess, Operation, Transition, CuttingParams
from app.agent import TechAgent


class ExcelTemplateFiller:
    """Заполнитель Excel шаблона Time Study"""
    
    def __init__(self, agent: TechAgent = None):
        self.agent = agent or TechAgent()
        # Получаем абсолютный путь к шаблону
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.template_path = os.path.join(
            current_dir, 
            "..", 
            "templates", 
            "Time Study_форма для заполнения.xlsx"
        )
        self.template_path = os.path.normpath(self.template_path)
    
    def fill_template(self, process: TechProcess, output_path: str) -> str:
        """Заполняет шаблон данными из технологического процесса"""
        
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Шаблон не найден: {self.template_path}")
        
        # Загружаем шаблон
        self.wb = load_workbook(self.template_path)
        self.ws = self.wb.active
        
        # Заполняем шапку
        self._fill_header(process)
        
        # Заполняем операции
        self._fill_operations(process)
        
        # Сохраняем
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        self.wb.save(output_path)
        
        return output_path
    
    def _fill_header(self, process: TechProcess):
        """Заполняет шапку документа (строки 3-16)"""
        ws = self.ws
        
        # Заказчик (C3)
        ws['C3'] = "ОАО «МЗШ»"
        
        # Модель оборудования (C4) - в шаблоне уже есть текст, заполняем значение
        ws['N4'] = "WINPOWER WPLF45D"
        
        # Система ЧПУ (C5)
        ws['N5'] = "SIEMENS 828D"
        
        # Макс. скорость шпинделя (C6)
        ws['N6'] = 2500
        
        # Макс. скорость приводн. инстр. (C7)
        ws['N7'] = 4500
        
        # Номер проекта (C8)
        ws['N8'] = ""
        
        # Номер детали (C9)
        ws['N9'] = process.drawing_number or ""
        
        # Наименование детали (C10)
        ws['N10'] = process.part_name
        
        # Материал (C11)
        ws['N11'] = process.material or ""
        
        # Твердость (C12)
        ws['N12'] = "156…207 HB"
        
        # Заготовка (C13)
        ws['N13'] = "Поковка Ø306.6 x 75 (22.1 кг)"
        
        # Коэффициент использования (C15)
        ws['N15'] = 0.85
    
    def _fill_operations(self, process: TechProcess):
        """Заполняет таблицу операций начиная со строки 20"""
        ws = self.ws
        
        # Очищаем существующие данные операций (строки 20-60)
        for row in range(20, 70):
            for col in range(1, 25):  # Колонки A-X
                ws.cell(row=row, column=col, value=None)
        
        # Начинаем заполнение (строка 20)
        current_row = 20
        
        for op in process.operations:
            # Заголовок операции (колонка B)
            ws[f'B{current_row}'] = f"Операция - {op.op_number}"
            current_row += 1
            
            # Установка заготовки (для первой операции)
            if op.op_number == "005":
                ws[f'B{current_row}'] = "Установка и закрепление заготовки в трехкулачковом патроне при помощи системы pick-up"
                ws.cell(row=current_row, column=24, value=15.0)  # Общее время (колонка X)
                current_row += 1
            
            # Переходы
            for trans in op.transitions:
                self._fill_transition_row(trans, current_row)
                current_row += 3  # Каждый инструмент занимает 3 строки
        
        # Снятие детали
        ws[f'B{current_row}'] = "Открепление и снятие готовой детали при помощи системы pick-up"
        ws.cell(row=current_row, column=24, value=15.0)
        current_row += 1
        
        # ИТОГО (формула уже есть в шаблоне)
        ws[f'B{current_row}'] = "ИТОГО:"
    
    def _fill_transition_row(self, trans: Transition, row: int):
        """Заполняет строку перехода (занимает 3 строки)"""
        ws = self.ws
        
        # Получаем данные инструмента по ID
        tool = self.agent.get_tool_by_id(trans.tool_id) if self.agent else None
        
        # № инструмента (колонка A, первая строка)
        tool_num = f"Т{trans.transition_number:03d}"
        ws.cell(row=row, column=1, value=tool_num)
        
        if tool:
            # Определяем тип инструмента и заполняем 3 строки
            desc_lower = tool['description'].lower()
            
            if "расточ" in desc_lower or "внутренн" in desc_lower:
                # Расточной инструмент
                ws.cell(row=row, column=2, value="Державка расточная Ø50")
                ws.cell(row=row, column=3, value="ID державка")
                ws.cell(row=row+1, column=2, value="Втулка переходная Ø50-32")
                ws.cell(row=row+1, column=3, value="Ø50-32")
                ws.cell(row=row+2, column=2, value="Резец расточной")
                ws.cell(row=row+2, column=3, value=tool['holder'])
                ws.cell(row=row+3, column=2, value="Пластина 80° (Ромб), R1.2")
                ws.cell(row=row+3, column=3, value=tool.get('full_designation', ''))
                
            elif "канавочн" in desc_lower:
                # Канавочный инструмент
                ws.cell(row=row, column=2, value="Державка")
                ws.cell(row=row, column=3, value="ID державка")
                ws.cell(row=row+1, column=2, value="Резец канавочный внутрен.")
                ws.cell(row=row+1, column=3, value=tool['holder'])
                ws.cell(row=row+2, column=2, value="Пластина b=2, R1")
                ws.cell(row=row+2, column=3, value=tool.get('full_designation', ''))
                
            elif "сверл" in desc_lower:
                # Сверло
                ws.cell(row=row, column=2, value="Осевой приводной блок")
                ws.cell(row=row, column=3, value="Axial Mill/ Drill unit")
                ws.cell(row=row+1, column=2, value="Цанга ER32-20")
                ws.cell(row=row+1, column=3, value="ER32 SPR 19-20")
                ws.cell(row=row+2, column=2, value="Сверло 5d со вставкой")
                ws.cell(row=row+2, column=3, value=tool['holder'])
                ws.cell(row=row+3, column=2, value="Вставка сверлильная Ø15.1")
                ws.cell(row=row+3, column=3, value=tool.get('full_designation', ''))
                
            elif "фасочн" in desc_lower or "фрез" in desc_lower:
                # Фасочная фреза
                ws.cell(row=row, column=2, value="Державка")
                ws.cell(row=row, column=3, value="ID державка")
                ws.cell(row=row+1, column=2, value="SL-ER STRAIGHT SHANK COLLET CHUCK")
                ws.cell(row=row+1, column=3, value="SL20-ER32-80L")
                ws.cell(row=row+2, column=2, value="Цанга ER32-20")
                ws.cell(row=row+2, column=3, value="ER32 SPR 19-20")
                ws.cell(row=row+3, column=2, value="Фреза фасочная")
                ws.cell(row=row+3, column=3, value=tool['holder'])
                ws.cell(row=row+4, column=2, value="Пластина")
                ws.cell(row=row+4, column=3, value=tool.get('full_designation', ''))
                
            else:
                # Проходной инструмент (по умолчанию)
                ws.cell(row=row, column=2, value="Державка")
                ws.cell(row=row, column=3, value="OD державка (перпендик. Z)")
                ws.cell(row=row+1, column=2, value="Резец проходной 32 x 32")
                ws.cell(row=row+1, column=3, value=tool['holder'])
                ws.cell(row=row+2, column=2, value=f"Пластина 80° ({tool.get('geometry', 'Ромб')})")
                ws.cell(row=row+2, column=3, value=tool.get('full_designation', ''))
        
        # Наименование перехода (колонка D, первая строка)
        ws.cell(row=row, column=4, value=trans.description)
        
        # Диаметры (колонки E, F)
        if trans.diameter:
            diam_match = re.search(r'Ø?(\d+\.?\d*)', trans.diameter)
            if diam_match:
                diam = float(diam_match.group(1))
                ws.cell(row=row, column=5, value=diam + 5)  # D max
                ws.cell(row=row, column=6, value=diam)      # D min
        
        # Режимы резания
        if trans.cutting_params:
            # V max/min (колонки G, H)
            v = trans.cutting_params.v_m_min or 0
            ws.cell(row=row, column=7, value=v)
            ws.cell(row=row, column=8, value=v)
            
            # n max/min (колонки I, J)
            n = trans.cutting_params.n_rpm or 0
            ws.cell(row=row, column=9, value=n)
            ws.cell(row=row, column=10, value=n)
            
            # Lрез (колонка K)
            ws.cell(row=row, column=11, value=30.0)
            
            # Lвсп (колонка L)
            ws.cell(row=row, column=12, value=5.0)
            
            # Припуск (колонка N)
            a = trans.cutting_params.t_mm or 2.0
            ws.cell(row=row, column=14, value=a)
            
            # Кол-во проход (колонка O)
            ws.cell(row=row, column=15, value=1)
            
            # Глубина резания t (колонка P)
            ws.cell(row=row, column=16, value=a)
            
            # Подача So (колонка R)
            s = trans.cutting_params.s_mm_rev or 0.3
            ws.cell(row=row, column=18, value=s)
            
            # Мощность P (колонка T)
            ws.cell(row=row, column=20, value=3.0)
            
            # Время смены инстр. (колонка U)
            ws.cell(row=row, column=21, value=3)
            
            # Время ускорен. перем. (колонка V)
            ws.cell(row=row, column=22, value=5)
            
            # Время резания (колонка W)
            cutting_time = self._calculate_cutting_time(trans)
            ws.cell(row=row, column=23, value=cutting_time)
            
            # Общее время (колонка X) - формула: =U+V+W
            ws.cell(row=row, column=24, value=f"=U{row}+V{row}+W{row}")
    
    def _calculate_cutting_time(self, trans: Transition) -> float:
        """Рассчитывает время резания в секундах"""
        if not trans.cutting_params:
            return 0.0
        
        l_cut = 30.0  # Примерная длина резания
        s = trans.cutting_params.s_mm_rev or 0.1
        n = trans.cutting_params.n_rpm or 100
        
        if s == 0 or n == 0:
            return 0.0
        
        time_sec = (l_cut / (s * n)) * 60
        return round(time_sec, 1)


class ExcelGenerator:
    """Основной генератор Excel файлов"""
    
    def __init__(self, agent: TechAgent = None):
        self.agent = agent or TechAgent()
        self.filler = ExcelTemplateFiller(self.agent)
    
    def generate(self, process: TechProcess, output_path: str) -> str:
        """Генерирует Excel файл Time Study"""
        return self.filler.fill_template(process, output_path)